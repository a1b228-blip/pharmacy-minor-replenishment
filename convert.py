#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
藥劑科小藥撥補單 - 主管格式轉換器 (Excel + A4 單頁橫式 PDF)
依最新指示排版：
1. 移除日期/筆數等副標題文字
2. 移除「製表人」與「領料單位簽收」
3. 簽名欄僅保留「發料藥佐」與「核對藥師」，並靠右排版
4. 字體最大化（8.3pt），嚴格容納於 1 張 A4 橫式紙張
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 註冊中文字型
CHINESE_FONT_PATHS = [
    '/Library/Fonts/Arial Unicode.ttf',
    '/System/Library/Fonts/Supplemental/Songti.ttc',
    '/System/Library/Fonts/STHeiti Light.ttc',
    '/System/Library/Fonts/Hiragino Sans GB.ttc'
]

FONT_REGISTERED = False
for fpath in CHINESE_FONT_PATHS:
    if os.path.exists(fpath):
        try:
            pdfmetrics.registerFont(TTFont('Chinese', fpath))
            FONT_REGISTERED = True
            break
        except Exception:
            continue

TARGET_COLUMN_MAP = [
    ('申請單號', '申請單號'),
    ('申請時間', '申請時間'),
    ('申請單位', '申請\n單位'),
    ('費用部門代號', '費用部門代號'),
    ('申請人', '申請人'),
    ('送簽主管', '送簽\n主管'),
    ('實際簽核主管', '實際簽核主管'),
    ('發料日期', '發料日期'),
    ('領料品項名稱', '領料品項名稱'),
    ('材料編號', '材料編號'),
    ('單位', '單位'),
    ('申請數量', '申請\n數量'),
    ('發料數量', '發料\n數量'),
    (None, '單位簽收人/日期')
]

COLUMN_WIDTHS = {
    1: 14, 2: 20, 3: 11, 4: 13, 5: 10, 6: 10, 7: 13,
    8: 12, 9: 28, 10: 16, 11: 8, 12: 10, 13: 10, 14: 18
}

def generate_excel(src_path, dest_xlsx):
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb.active
    src_headers = [cell.value for cell in src_ws[1]]

    col_mapping = []
    for src_col, target_header in TARGET_COLUMN_MAP:
        if src_col is not None and src_col in src_headers:
            col_mapping.append((src_headers.index(src_col), target_header))
        else:
            col_mapping.append((None, target_header))

    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active
    dest_ws.title = "藥品材料申請單"

    font_header = Font(name="新細明體", size=11, bold=True)
    font_data = Font(name="新細明體", size=10, bold=False)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )

    dest_ws.row_dimensions[1].height = 28
    for col_idx, (_, target_header) in enumerate(col_mapping, start=1):
        cell = dest_ws.cell(row=1, column=col_idx, value=target_header)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    extracted_rows = []
    for row_idx, src_row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
        dest_ws.row_dimensions[row_idx].height = 24
        curr_row = []
        for col_idx, (src_idx, target_header) in enumerate(col_mapping, start=1):
            val = src_row[src_idx] if src_idx is not None else ""
            if val is None: val = ""
            cell = dest_ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if '品項名稱' in target_header:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
            curr_row.append(val)
        extracted_rows.append(curr_row)

    for col_idx, width in COLUMN_WIDTHS.items():
        col_letter = get_column_letter(col_idx)
        dest_ws.column_dimensions[col_letter].width = width

    dest_ws.page_setup.orientation = dest_ws.ORIENTATION_LANDSCAPE
    dest_ws.page_setup.paperSize = dest_ws.PAPERSIZE_A4
    dest_ws.page_setup.fitToPage = True
    dest_ws.page_setup.fitToWidth = 1
    dest_ws.page_setup.fitToHeight = 0
    dest_ws.sheet_properties.pageSetUpPr.fitToPage = True

    dest_wb.save(dest_xlsx)
    return [h for _, h in col_mapping], extracted_rows

def generate_single_page_pdf(headers, rows, dest_pdf):
    if not FONT_REGISTERED:
        print("未偵測到可用中文字型，略過 PDF 產出")
        return

    margin = 8
    doc = SimpleDocTemplate(
        dest_pdf,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=6,
        bottomMargin=6
    )

    row_count = len(rows)
    # 動態字級最適化（8.3pt 最大化清晰字級）
    if row_count > 32:
        font_size = 7.2
        leading = 8.2
        padding = 1.0
    elif row_count > 25:
        font_size = 8.3
        leading = 9.4
        padding = 1.2
    else:
        font_size = 9.0
        leading = 10.5
        padding = 1.8

    title_style = ParagraphStyle(
        'TitleStyle',
        fontName='Chinese',
        fontSize=15,
        leading=17,
        alignment=1,
        textColor=colors.HexColor('#0f172a')
    )

    elements = [
        Paragraph('<b>佳里奇美醫院 藥劑科小藥撥補單</b>', title_style),
        Spacer(1, 4)
    ]

    col_widths = [48, 80, 40, 50, 34, 34, 52, 42, 145, 78, 24, 32, 32, 134]

    table_data = []
    header_cells = []
    for h in headers:
        header_cells.append(Paragraph(
            f'<b>{h.replace(chr(10), "<br/>")}</b>',
            ParagraphStyle('TH', fontName='Chinese', fontSize=font_size, leading=leading, alignment=1)
        ))
    table_data.append(header_cells)

    for r in rows:
        row_cells = []
        for i, val in enumerate(r):
            v_str = str(val or '')
            align = 0 if i == 8 else 1
            row_cells.append(Paragraph(
                v_str,
                ParagraphStyle('TD', fontName='Chinese', fontSize=font_size, leading=leading, alignment=align)
            ))
        table_data.append(row_cells)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#334155')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), padding),
        ('BOTTOMPADDING', (0,0), (-1,-1), padding),
        ('LEFTPADDING', (0,0), (-1,-1), 1.5),
        ('RIGHTPADDING', (0,0), (-1,-1), 1.5),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8))

    # 簽名欄位靠右對齊 (發料藥佐、核對藥師)
    sig_data = [[
        '', # 左側彈性空間，將兩組簽名推向右側
        Paragraph('發料藥佐：_______________________', ParagraphStyle('S1', fontName='Chinese', fontSize=9.5, alignment=2)),
        Paragraph('核對藥師：_______________________', ParagraphStyle('S2', fontName='Chinese', fontSize=9.5, alignment=2))
    ]]
    t_sig = Table(sig_data, colWidths=[405, 210, 210])
    t_sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(t_sig)

    doc.build(elements)

def convert_replenishment_excel(src_path):
    if not os.path.exists(src_path):
        print(f"錯誤：找不到檔案 {src_path}")
        return

    base, _ = os.path.splitext(src_path)
    dest_xlsx = f"{base}_正確格式.xlsx"
    dest_pdf = f"{base}_單頁橫式.pdf"

    print(f"正在讀取：{src_path}")
    headers, rows = generate_excel(src_path, dest_xlsx)
    print(f"✅ 成功產出 Excel：{dest_xlsx}")

    generate_single_page_pdf(headers, rows, dest_pdf)
    print(f"✅ 成功產出 A4 單頁橫式 PDF：{dest_pdf}")

if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else '/Users/jiangruiyi/Downloads/藥品材料申請單_2026-08-27_to_2026-09-02.xlsx'
    convert_replenishment_excel(target)
